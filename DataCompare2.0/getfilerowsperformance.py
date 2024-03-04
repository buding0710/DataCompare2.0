#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2023/4/13 17:42
# @Author   : buding
# @Project  : pythonProject
# @File     : getfilerowsperformance.py
# @Software : PyCharm
import csv
import sys
import time
from functools import wraps

import openpyxl
import pandas as pd
import xlrd


def funcName(func_name):
    def runTime(func):
        def warp(*args,**kwargs):
            start = time.time()
            func(*args,**kwargs)
            end = time.time()
            print(func_name + "耗时：" + str((end - start) * 1000))
        return warp
    return runTime


class GetRows:
    def __init__(self,path):
        self.path = path

    @funcName('lenReadLines')
    def lenReadLines(self):
        print(len(open(self.path, 'rb').readlines()))

    @funcName('countLines')
    def countLines(self):
        lines = 0
        for _ in open(self.path):
            lines += 1
        print(lines)

    @funcName('sumLines')
    def sumLines(self):
        print(sum(1 for _ in open(self.path)))

    @funcName('countEnumerate')
    def countEnumerate(self):
        # 文件比较大
        count = 0
        for count, line in enumerate(open(self.path, 'rb')):
            count += 1
        print(count)

    # with open(file_name) as f:
    #     for count, _ in enumerate(f, 1):
    #         pass
    # return count

    @funcName('partCount')
    def partCount(self):
        with open(self.path, 'rb') as f:
            count = 0
            part_size = 1024 * 1024
            part = f.read(part_size)
            while part:
                count += part.count(b'\n')
                part = f.read(part_size)
            print(count)

    @funcName('wc_count')
    def wc_count(self):
        import subprocess
        # out = subprocess.getoutput("wc -l %s" % path) #linux
        str = f'type {self.path} | find /v /c ""' #windows
        out = subprocess.getoutput(str)
        print(int(out.split()[0]))

    @funcName('partial_count')
    def partial_count(self):
        from functools import partial
        buffer = 1024 * 1024
        with open(self.path) as f:
            print(sum(x.count('\n') for x in iter(partial(f.read, buffer), '')))

    @funcName('iter_count')
    def iter_count(self):
        from itertools import (takewhile, repeat)
        buffer = 1024 * 1024
        with open(self.path) as f:
            buf_gen = takewhile(lambda x: x, (f.read(buffer) for _ in repeat(None)))
            print(sum(buf.count('\n') for buf in buf_gen))

    @funcName('df_read_len')
    def df_read_len(self):
        df = pd.read_csv(self.path)
        print(len(df))

    @funcName('df_read_shape')
    def df_read_shape(self):
        df = pd.read_csv(self.path)
        print(df.shape[0])

    @funcName('countLF')
    def countLF(self):
        with open(self.path, 'rb') as f:
            count = 0
            while True:
                data = f.read(0x400000)
                if not data:
                    break
                count += data.count(b'\n')
        print(count)


class GetColumns():

    def __init__(self,path):
        self.path = path

    @funcName('readerRows')
    def readerRows(self):
        with open(self.path, "r",encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile)
            rows = [row for row in reader]
        print(len(rows[0]))

    @funcName('dictReaderCount')
    def dictReaderCount(self):
        with open(self.path, "rt",encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            column = [row for row in reader]
        print(len(column[0]))

    @funcName('dataframe1')
    def dataframe1(self):
        df = pd.read_csv(self.path)
        print(df.shape[1])

    @funcName('dataframe2')
    def dataframe2(self):
        df = pd.read_csv(self.path)
        print(len(df.columns))


class GetExcelRows:
    def __init__(self,path):
        self.path = path

    @funcName('openpyxlRows')
    def openpyxlRows(self):#不支持XLS格式
        # 打开Excel文件并选择工作表
        wb = openpyxl.load_workbook(self.path)
        sheets = wb.get_sheet_names()
        sheet = wb[sheets[0]]
        print(sheet.max_row)
        print(sheet.max_column)

    @funcName('readExcelRows')
    def readExcelRows(self):#支持xls和xlsx的格式
        #read_excel()用来读取excel文件，记得加文件后缀
        data = pd.read_excel(self.path)
        print(data.shape[0])   #打印显示表格的属性，几行几列
        print(data.shape[1])

    @funcName('xlrdRows')
    def xlrdRows(self):#不支持xlsx格式
        data = xlrd.open_workbook(self.path)
        table = data.sheets()[0]
        print(table.nrows) # 获取该sheet中的有效行数
        print(table.ncols)  # 获取列表的有效列数

    @funcName('lenReadLines')
    def lenReadLines(self):
        print(len(open(self.path, 'rb').readlines()))

    @funcName('countLines')
    def countLines(self):
        lines = 0
        for _ in open(self.path,'rb'):
            lines += 1
        print(lines)

    @funcName('sumLines')
    def sumLines(self):
        print(sum(1 for _ in open(self.path,'rb')))

    @funcName('countEnumerate')
    def countEnumerate(self):
        # 文件比较大
        count = 0
        for count, line in enumerate(open(self.path, 'rb')):
            count += 1
        print(count)

    @funcName('wc_count')
    def wc_count(self):
        import subprocess
        # out = subprocess.getoutput("wc -l %s" % path) #linux
        str = f'type {self.path} | find /v /c ""'  # windows
        out = subprocess.getoutput(str)
        print(int(out.split()[0]))

    @funcName('countLF')
    def countLF(self):
        with open(self.path, 'rb') as f:
            count = 0
            while True:
                data = f.read(0x400000)
                if not data:
                    break
                count += data.count(b'\n')
        print(count)

class ReadFileConvertToDF():
    def __init__(self,path,skip_rows,nrows):
        self.path = path
        self.skip_rows = skip_rows
        self.nrows = nrows

    @funcName('CsvOrTxtData')
    def CsvOrTxtData(self):#read_csv支持csv和txt的格式
        df = pd.read_csv(self.path,skiprows=self.skip_rows,nrows=self.nrows,)
        print(df)

    @funcName('openFileConvertToDF')
    def openFileConvertToDF(self):#with open()仅支持csv和txt的格式
        rows = []
        with open(path,"r",encoding="utf-8") as f:
            for row in f:
                rows.append(row.rstrip('\n').split(','))
        s = self.skip_rows + 1 if self.skip_rows == 0 else self.skip_rows
        n = self.nrows + self.skip_rows
        df = pd.DataFrame(rows[s:n+1],columns=rows[0])
        print(df)

if __name__ == '__main__':
    path = r"D:\pythonProject\DataCompare\test.csv"
    # path = "mapping1.txt"
    # GetColumns(path).readerRows()
    # GetColumns(path).dictReaderCount()
    # GetColumns(path).dataframe1()
    # GetColumns(path).dataframe2()

    GetExcelRows(path).readExcelRows()
    # GetExcelRows(path).openpyxlRows()
    GetExcelRows(path).xlrdRows()
    GetExcelRows(path).lenReadLines()
    GetExcelRows(path).countLines()
    GetExcelRows(path).sumLines()
    GetExcelRows(path).countEnumerate()
    GetExcelRows(path).wc_count()
    GetExcelRows(path).countLF()

    # GetRows(path).lenReadLines()
    # GetRows(path).countLines()
    # GetRows(path).sumLines()
    # GetRows(path).countEnumerate()
    # GetRows(path).partCount()
    # GetRows(path).wc_count()
    # GetRows(path).partial_count()
    # GetRows(path).iter_count()
    # GetRows(path).df_read_len()
    # GetRows(path).df_read_shape()
    # GetRows(path).countLF()

    # ReadFileConvertToDF(path,5,5).openFileConvertToDF()
    # ReadFileConvertToDF(path, 5, 5).CsvOrTxtData()
    # ReadFileConvertToDF(path, 0, 5).XlsOrXlsxData()